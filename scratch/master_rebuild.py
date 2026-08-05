"""
Master script: 
1. Reads Excel for real product specs + pricing + SEO meta
2. Rebuilds all product pages with REAL data
3. Applies unified mega menu nav to ALL pages
4. Fixes all links
5. Updates index.html
"""
import openpyxl, os, re, json

PROJ = r'c:\Users\harsh\OneDrive\Desktop\IIMBx live project krishtech site agents latest'
VARIANTS = os.path.join(PROJ, 'variants')
EXCEL = os.path.join(PROJ, 'KIPL_Product Specification_FM-DSN-05 (1)krishitek.xlsx')

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ===== PARSE SPECS =====
ws = wb['Specs with Features']
products = {}
current_product = None
current_specs = []
current_features = []
current_model = None

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6, values_only=True):
    vals = [str(v).strip() if v else '' for v in row]
    
    # Detect product header (numbered items like "1.0", "7.0" in col A)
    if vals[0] and re.match(r'^\d+\.0$', vals[0]):
        if current_product and (current_specs or current_features):
            products[current_product] = {'model': current_model, 'specs': current_specs, 'features': current_features}
        current_product = vals[1].strip()
        current_specs = []
        current_features = []
        current_model = None
    elif current_product and vals[0] == '':
        # Spec row: label in col B (idx 1), value in col C (idx 2)
        label = vals[1].strip()
        value = vals[2].strip()
        if label and value and label not in ('', 'None') and label != 'Category':
            if 'Category' in label:
                continue
            if label == 'Model':
                current_model = value
            if label not in ('Dimensions', 'Packaging', 'Spare Parts Price List'):
                current_specs.append((label, value))
        # Feature: number in col D (idx 3), text in col E (idx 4)
        feat_text = vals[4].strip() if len(vals) > 4 else ''
        if vals[3] and re.match(r'^\d+\.0$', vals[3]) and feat_text:
            # Take the part before the first colon as the feature title
            if ':' in feat_text:
                feat_title = feat_text.split(':')[0].strip()
            else:
                feat_title = feat_text[:80]
            current_features.append(feat_title)

if current_product and (current_specs or current_features):
    products[current_product] = {'model': current_model, 'specs': current_specs, 'features': current_features}

# ===== PARSE PRICING =====
ws2 = wb['Product pricing']
pricing = {}
for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, max_col=5, values_only=True):
    vals = [str(v).strip() if v else '' for v in row]
    if vals[1]:
        name = vals[1]
        mrp = vals[3]
        web = vals[4]
        pricing[name] = {'mrp': mrp, 'web_price': web, 'type': vals[2]}

print("=== PRODUCTS PARSED ===")
for name, data in products.items():
    print(f"  {name}: model={data['model']}, specs={len(data['specs'])}, features={len(data['features'])}")

print(f"\n=== PRICING PARSED: {len(pricing)} items ===")
for name in pricing:
    print(f"  {name}: Rs.{pricing[name]['web_price']}")

# ===== MAP products to filenames =====
PRODUCT_MAP = {
    'SELF PROPELLED POWER REAPER': {
        'file': 'reaptek_ki_120_product_page.html',
        'category': 'Self Propelled Reapers',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/Self-Propelled-Power-Reape-300x300.jpg',
        'price_key': 'Reaptek KI120 (with Honda Engine)',
    },
    'POWER REAPER CUM WEEDER (Dual function)': {
        'file': 'reaptek_combine_4_8wp_product_page.html',
        'category': 'Power Reapers',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/Power-Reaper-Cum-Weeder-Dual-Function-300x300.jpg',
        'price_key': 'Reaptek Combine (Dual Function) (with Honda engine)',
    },
    'POWER WEEDER 5.5WP': {
        'file': 'power_weeder_5_5wp_product_page.html',
        'category': 'Power Weeders',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/Power-Weeder-5.5WP-300x300.jpg',
        'price_key': 'Powertek 5.5WP (with Honda Engine)',
    },
    'POWER WEEDER 7CR': {
        'file': 'power_weeder_7cr_product_page.html',
        'category': 'Power Weeders',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/power-weeder-7cr-300x300.jpg',
        'price_key': 'Powertek 7CR',
    },
    'BACK ROTARY POWER WEEDER 7BR': {
        'file': 'back_rotary_power_weeder_7br_product_page.html',
        'category': 'Power Weeders',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/Back-rotary-power-weeder-7BR-300x300.jpg',
        'price_key': 'Powertek 7BR (with standard blade set)',
    },
    'FRONT ROTARY POWER WEEDER 7FR': {
        'file': 'front_rotary_power_weeder_7fr_product_page.html',
        'category': 'Power Weeders',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/Front-rotary-power-weeder-300x300.jpg',
        'price_key': 'Powertek 7FR',
    },
    'MINI POWER TILLER 3WP': {
        'file': 'mini_tiller_powertek_3wp_product_page.html',
        'category': 'Mini Tillers',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/Mini-Tiller-Powertek-3WP-300x300.jpg',
        'price_key': 'Powertek 3WP mini tiller',
    },
    '3FT REAPER ATTACHMENT FOR POWER WEEDER': {
        'file': 'reaptek_3_product_page.html',
        'category': 'Reaper Attachments',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/3ft-reaper-attachment-300x300.jpg',
        'price_key': 'Reaptek 3 + Powertek 7CR',
    },
    'REAPER ATTCHMENT FOR SWARAJ CODE (Self Propelled Tool Bar)': {
        'file': 'reaptek_c4_c5_product_page.html',
        'category': 'Tractor Attachments',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/Reaptek-C4-C5-300x300.jpg',
        'price_key': 'Reaptek C5 (5ft cutter bar)',
    },
    'REAPER ATTCHMENT FOR POWER TILLER': {
        'file': 'reaptek_pt4_pt5_product_page.html',
        'category': 'Tractor Attachments',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/power-tiller-reaper-300x300.jpg',
        'price_key': 'Reaptek PT5 (5ft cutter bar)',
    },
    'TRACTOR OPERATED HYDRAULIC REAPER ( Hydraulic vertical Conveyor Reaper)': {
        'file': 'reaptek_t7_t6_t5_t4_product_page.html',
        'category': 'Tractor Reapers',
        'image': 'https://www.krishitek.com/wp-content/uploads/2024/07/Tractor-Mounted-Reaper-300x300.jpg',
        'price_key': 'Reaptek T7 (7ft cutter bar)',
    },
}

# ===== MEGA MENU CSS =====
MEGA_CSS = """
        /* Mega Menu */
        .mega-trigger { position: static; display: flex; align-items: center; gap: 5px; padding: 10px 0; }
        .mega-trigger i.fa-plus { font-size: 10px; transition: transform 0.3s; }
        .mega-trigger:hover i.fa-plus { transform: rotate(45deg); color: var(--logo-red); }
        .mega-menu { display: none; position: absolute; top: 100%; left: 0; width: 100%; background: var(--white); box-shadow: 0 20px 60px rgba(0,0,0,0.15); border-top: 3px solid var(--logo-red); z-index: 999; }
        .mega-trigger::after { content: ''; position: absolute; top: 100%; left: 0; width: 100%; height: 20px; background: transparent; }
        .mega-trigger:hover .mega-menu, .mega-menu:hover { display: flex; }
        .mega-col-1 { width: 240px; flex-shrink: 0; background: var(--gray-light); padding: 32px 0; border-right: 1px solid var(--gray-border); }
        .mega-col-1 a { display: flex; align-items: center; justify-content: space-between; padding: 14px 28px; font-family: 'Barlow', sans-serif; font-weight: 700; font-size: 15px; text-transform: uppercase; color: var(--charcoal); text-decoration: none; transition: all 0.15s; letter-spacing: 0.5px; }
        .mega-col-1 a:hover, .mega-col-1 a.active { color: var(--logo-red); background: var(--white); border-left: 3px solid var(--logo-red); }
        .mega-col-1 a i { font-size: 11px; color: var(--logo-red); opacity: 0; transition: opacity 0.2s; }
        .mega-col-1 a.active i { opacity: 1; }
        .mega-panels { flex: 1; display: flex; }
        .mega-panel { display: none; width: 100%; }
        .mega-panel.active { display: flex; }
        .mega-col-2 { width: 250px; flex-shrink: 0; padding: 32px 24px; border-right: 1px solid var(--gray-border); }
        .mega-col-2 a { display: flex; align-items: center; justify-content: space-between; padding: 11px 14px; font-family: 'Barlow', sans-serif; font-weight: 600; font-size: 14px; text-transform: uppercase; color: var(--charcoal); text-decoration: none; border-radius: 6px; transition: all 0.15s; letter-spacing: 0.3px; }
        .mega-col-2 a:hover, .mega-col-2 a.active { color: var(--logo-red); background: rgba(192,0,0,0.05); }
        .mega-col-3 { flex: 1; padding: 32px 40px; }
        .mega-col-3 h4 { font-family: 'Barlow', sans-serif; font-size: 14px; font-weight: 700; text-transform: uppercase; color: var(--ink); margin-bottom: 18px; letter-spacing: 1.5px; }
        .mega-products-list { display: grid; grid-template-columns: 1fr 1fr; gap: 4px 40px; }
        .mega-products-list a { font-size: 14px; font-family: 'Inter', sans-serif; font-weight: 500; color: var(--charcoal); text-decoration: none; padding: 8px 0; display: block; line-height: 1.5; transition: color 0.15s; }
        .mega-products-list a::before { content: '\\2022  '; color: var(--logo-red); font-weight: bold; }
        .mega-products-list a:hover { color: var(--logo-red); }
        .mega-products-list .product-sub { font-size: 12px; color: #999; display: block; margin-top: 1px; padding-left: 14px; font-style: italic; font-weight: 400; }
        @media (max-width: 900px) { .mega-menu { display: none !important; } }
"""

# ===== NAV HTML =====
NAV_HTML = '''<div class="utility-bar">
<a class="dealer-btn" href="contact_v5_premium.html">Dealer Inquiry Form</a>
<div class="lang-dropdown">
<img alt="EN" src="https://flagcdn.com/w20/gb.png" width="16"/> English <i class="fas fa-caret-down"></i>
<div class="lang-menu">
<a href="#"><img alt="IN" src="https://flagcdn.com/w20/in.png" width="16"/> Hindi</a>
<a href="#"><img alt="IN" src="https://flagcdn.com/w20/in.png" width="16"/> Gujarati</a>
<a href="#"><img alt="IN" src="https://flagcdn.com/w20/in.png" width="16"/> Marathi</a>
<a href="#"><img alt="IN" src="https://flagcdn.com/w20/in.png" width="16"/> Tamil</a>
</div>
</div>
<div class="social-links">
<a class="bg-fb" href="#"><i class="fab fa-facebook-f"></i></a>
<a class="bg-in" href="#"><i class="fab fa-linkedin-in"></i></a>
<a class="bg-ig" href="#"><i class="fab fa-instagram"></i></a>
<a class="bg-phone" href="#"><i class="fas fa-phone"></i></a>
<a class="bg-map" href="#"><i class="fas fa-map-marker-alt"></i></a>
</div>
</div>
<nav>
<a class="logo" href="homepage_v4_stitch.html">
<img alt="KrishiTek Logo" onerror="this.src='https://via.placeholder.com/150x50/FFFFFF/C00000?text=KRISHITEK'" src="https://www.krishitek.com/wp-content/uploads/2022/11/LOGO.png"/>
KRISHITEK
</a>
<div class="nav-links">
<a href="about_v1.html">ABOUT</a>
<div class="mega-trigger">
PRODUCTS <i class="fas fa-plus" style="font-size:10px;"></i>
<div class="mega-menu">
<div class="mega-col-1">
<a class="active" href="product_catalog_v1.html" data-target="agri">AGRI MACHINERY <i class="fas fa-chevron-right"></i></a>
<a href="spare_parts_v1.html" data-target="spare">SPARE PARTS <i class="fas fa-chevron-right"></i></a>
<a href="services_v1.html" data-target="services">SERVICES &amp; SUPPORT <i class="fas fa-chevron-right"></i></a>
</div>
<div class="mega-panels">
  <!-- Agri Panel -->
  <div class="mega-panel active" id="mega-agri">
    <div class="mega-col-2">
    <a class="active" href="#">REAPERS &amp; HARVESTERS</a>
    <a href="#">POWER WEEDERS</a>
    <a href="#">MINI TILLERS</a>
    <a href="#">TRACTOR ATTACHMENTS</a>
    </div>
    <div class="mega-col-3">
    <h4>PRODUCTS</h4>
    <div class="mega-products-list">
    <a href="reaptek_ki_120_product_page.html">Self Propelled Power Reaper<span class="product-sub">Reaptek Ki-120</span></a>
    <a href="reaptek_combine_4_8wp_product_page.html">Power Reaper Cum Weeder<span class="product-sub">Reaptek Combine 4.8WP</span></a>
    <a href="power_weeder_5_5wp_product_page.html">Power Weeder 5.5WP<span class="product-sub">Honda GX200 Engine</span></a>
    <a href="power_weeder_7cr_product_page.html">Power Weeder 7CR<span class="product-sub">7HP Chain Drive</span></a>
    <a href="back_rotary_power_weeder_7br_product_page.html">Back Rotary Weeder 7BR<span class="product-sub">Patented Rotor</span></a>
    <a href="front_rotary_power_weeder_7fr_product_page.html">Front Rotary Weeder 7FR<span class="product-sub">Paddy Puddling</span></a>
    <a href="mini_tiller_powertek_3wp_product_page.html">Mini Tiller 3WP<span class="product-sub">Compact Powertek</span></a>
    <a href="reaptek_3_product_page.html">Reaper Attachment 3ft<span class="product-sub">For Power Weeder</span></a>
    </div>
    </div>
  </div>
  <!-- Spare Parts Panel -->
  <div class="mega-panel" id="mega-spare">
    <div class="mega-col-2">
    <a class="active" href="#">REAPER COMPONENTS</a>
    <a href="#">WEEDER COMPONENTS</a>
    <a href="#">ENGINE PARTS</a>
    <a href="#">ACCESSORIES</a>
    </div>
    <div class="mega-col-3">
    <h4>POPULAR SPARES</h4>
    <div class="mega-products-list">
    <a href="spare_clutch_lever_product_page.html">Clutch Lever<span class="product-sub">For Power Reapers</span></a>
    <a href="spare_big_blade_product_page.html">Big Blade<span class="product-sub">Cutting Mechanism</span></a>
    <a href="spare_chain_cover_assy_product_page.html">Chain Cover Assembly<span class="product-sub">Protection</span></a>
    <a href="spare_harvesting_gear_box_assy_product_page.html">Harvesting Gear Box<span class="product-sub">Power Transmission</span></a>
    <a href="spare_rotavator_blade_product_page.html">Rotavator Blade<span class="product-sub">For Weeders</span></a>
    <a href="spare_recoil_starter_product_page.html">Recoil Starter<span class="product-sub">Engine Component</span></a>
    </div>
    </div>
  </div>
  <!-- Services Panel -->
  <div class="mega-panel" id="mega-services">
    <div class="mega-col-2">
    <a class="active" href="#">WARRANTY REGISTRATION</a>
    <a href="#">SERVICE REQUEST</a>
    <a href="#">TUTORIALS &amp; GUIDES</a>
    <a href="#">DEALERSHIP</a>
    </div>
    <div class="mega-col-3">
    <h4>SUPPORT LINKS</h4>
    <div class="mega-products-list">
    <a href="services_v1.html">Register Warranty<span class="product-sub">Protect your investment</span></a>
    <a href="services_v1.html">Book a Service<span class="product-sub">Expert maintenance</span></a>
    <a href="media_v1_stitch.html">Video Tutorials<span class="product-sub">Operation guides</span></a>
    <a href="dealership_v1.html">Become a Dealer<span class="product-sub">Join our network</span></a>
    <a href="contact_v5_premium.html">Contact Support<span class="product-sub">Get in touch</span></a>
    <a href="blog_v1.html">Knowledge Base<span class="product-sub">Farming tips</span></a>
    </div>
    </div>
  </div>
</div>
</div>
</div>
<a href="dealership_v1.html">FIND A DEALER</a>
<a href="blog_v1.html">BLOG</a>
<a href="media_v1_stitch.html">MEDIA</a>
<a href="spare_parts_v1.html">SPARE PARTS</a>
<a href="contact_v5_premium.html">CONTACT</a>
</div>
<div class="menu-toggle">&#9776;</div>
</nav>
<script>
  // Mega Menu Hover Logic
  document.addEventListener('DOMContentLoaded', () => {
    const tabs = document.querySelectorAll('.mega-col-1 a[data-target]');
    const panels = document.querySelectorAll('.mega-panel');
    
    tabs.forEach(tab => {
      tab.addEventListener('mouseenter', () => {
        // Remove active class from all tabs and panels
        tabs.forEach(t => t.classList.remove('active'));
        panels.forEach(p => p.classList.remove('active'));
        
        // Add active class to current tab and panel
        tab.classList.add('active');
        const targetId = 'mega-' + tab.getAttribute('data-target');
        const targetPanel = document.getElementById(targetId);
        if(targetPanel) targetPanel.classList.add('active');
      });
    });
  });
</script>'''

# ===== PRODUCT PAGE TEMPLATE =====
def build_product_page(product_name, pdata, pmap):
    title = product_name.title()
    model = pdata.get('model', title)
    specs = pdata.get('specs', [])
    features = pdata.get('features', [])
    category = pmap.get('category', 'Agriculture Machinery')
    image = pmap.get('image', '')
    
    price_key = pmap.get('price_key', '')
    price_info = pricing.get(price_key, {})
    web_price = price_info.get('web_price', '')
    price_display = f'₹{int(float(web_price)):,}' if web_price and web_price != 'None' else 'Get Quote'
    
    specs_html = ''
    for label, value in specs[:10]:
        specs_html += f'<div class="spec-item"><span class="spec-label">{label}</span><span class="spec-value">{value}</span></div>\n'
    
    features_html = ''
    for feat in features[:6]:
        features_html += f'<li>{feat}</li>\n'
    if not features_html:
        features_html = '<li>Made in India — 100% indigenous manufacturing</li>\n<li>Genuine KrishiTek quality with warranty</li>\n<li>Easy availability of spare parts</li>\n'

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta content="width=device-width, initial-scale=1.0" name="viewport"/>
<title>KrishiTek - {title}</title>
<link href="https://fonts.googleapis.com/css2?family=Barlow:wght@400;500;600;700&amp;family=Inter:wght@400;500;600&amp;display=swap" rel="stylesheet"/>
<link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet"/>
<style>
        :root {{ --brand-yellow: #FDB913; --logo-red: #C00000; --charcoal: #333333; --ink: #1A1A1A; --white: #FFFFFF; --yellow-tint: #FEF3D0; --gray-light: #F7F9FA; --gray-border: #EAEAEA; }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Inter', sans-serif; color: var(--charcoal); background-color: var(--white); line-height: 1.6; -webkit-font-smoothing: antialiased; }}
        h1, h2, h3, h4, h5, h6, .btn {{ font-family: 'Barlow', sans-serif; }}
        .utility-bar {{ position: relative; z-index: 1001; background: var(--white); padding: 8px 5%; display: flex; justify-content: flex-end; align-items: center; gap: 24px; font-size: 13px; border-bottom: 1px solid var(--gray-border); }}
        .dealer-btn {{ background: #28a745; color: var(--white); padding: 6px 16px; border-radius: 4px; text-decoration: none; font-weight: 600; }}
        .lang-dropdown {{ position: relative; cursor: pointer; font-weight: 600; display: flex; align-items: center; gap: 6px; color: var(--ink); }}
        .lang-dropdown:hover .lang-menu {{ display: block; }}
        .lang-menu {{ display: none; position: absolute; top: 100%; right: 0; background: var(--white); box-shadow: 0 5px 15px rgba(0,0,0,0.1); border-radius: 4px; padding: 10px 0; z-index: 1001; min-width: 140px; border: 1px solid var(--gray-border); }}
        .lang-menu a {{ display: flex; padding: 8px 16px; color: var(--charcoal); text-decoration: none; font-weight: 500; align-items: center; gap: 8px; }}
        .lang-menu a:hover {{ background: var(--gray-light); color: var(--logo-red); }}
        .social-links {{ display: flex; gap: 4px; }}
        .social-links a {{ display: flex; align-items: center; justify-content: center; width: 26px; height: 26px; color: var(--white); text-decoration: none; border-radius: 2px; font-size: 13px; }}
        .bg-fb {{ background: #3b5998; }} .bg-in {{ background: #007bb5; }} .bg-ig {{ background: #e1306c; }} .bg-phone, .bg-map {{ background: var(--logo-red); }}
        nav {{ display: flex; justify-content: space-between; align-items: center; padding: 1.2rem 5%; background: rgba(255,255,255,0.95); backdrop-filter: blur(10px); position: sticky; top: 0; z-index: 1000; border-bottom: 1px solid var(--gray-border); box-shadow: 0 4px 20px rgba(0,0,0,0.03); position: relative; }}
        .logo {{ font-weight: 700; font-size: 24px; color: var(--logo-red); text-decoration: none; letter-spacing: 0.5px; display: flex; align-items: center; gap: 10px; }}
        .logo img {{ height: 40px; width: auto; object-fit: contain; }}
        .nav-links {{ display: flex; gap: 1.5rem; flex-wrap: wrap; justify-content: center; align-items: center; }}
        .nav-links > a, .nav-links > .mega-trigger {{ text-decoration: none; color: var(--charcoal); font-weight: 600; font-size: 14px; transition: color 0.2s; cursor: pointer; }}
        .nav-links > a:hover, .nav-links > .mega-trigger:hover {{ color: var(--logo-red); }}
        .btn {{ background-color: var(--logo-red); color: var(--white); padding: 12px 28px; border: none; border-radius: 6px; font-weight: 600; cursor: pointer; text-decoration: none; transition: transform 0.2s, box-shadow 0.2s; }}
        .btn:hover {{ transform: translateY(-2px); box-shadow: 0 6px 15px rgba(192, 0, 0, 0.2); }}
        .menu-toggle {{ display: none; font-size: 24px; cursor: pointer; }}
{MEGA_CSS}
        .product-layout {{ display: flex; padding: 80px 5%; gap: 60px; max-width: 1400px; margin: 0 auto; flex-wrap: wrap; }}
        .product-gallery {{ flex: 1; min-width: 300px; }}
        .main-image {{ width: 100%; min-height: 400px; background: var(--gray-light); border-radius: 16px; display: flex; align-items: center; justify-content: center; overflow: hidden; border: 1px solid var(--gray-border); padding: 20px; }}
        .main-image img {{ width: 100%; height: auto; object-fit: contain; mix-blend-mode: multiply; transition: transform 0.5s ease; }}
        .main-image:hover img {{ transform: scale(1.05); }}
        .product-details {{ flex: 1; min-width: 300px; display: flex; flex-direction: column; justify-content: center; }}
        .product-category {{ color: var(--logo-red); font-family: 'Barlow'; font-weight: 700; font-size: 14px; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 12px; }}
        .product-title {{ font-size: 42px; color: var(--ink); margin-bottom: 20px; line-height: 1.1; letter-spacing: -0.5px; }}
        .product-price {{ font-size: 28px; color: var(--charcoal); font-weight: bold; margin-bottom: 30px; font-family: 'Barlow'; }}
        .product-price .mrp {{ text-decoration: line-through; color: #999; font-size: 18px; margin-left: 12px; font-weight: 500; }}
        .product-description {{ font-size: 17px; color: #555; margin-bottom: 40px; line-height: 1.8; }}
        .feature-list {{ list-style: none; margin-bottom: 50px; }}
        .feature-list li {{ padding-left: 36px; position: relative; margin-bottom: 14px; font-weight: 500; font-size: 15px; }}
        .feature-list li::before {{ content: '\\2713'; color: var(--white); background: var(--brand-yellow); width: 24px; height: 24px; border-radius: 50%; display: flex; align-items: center; justify-content: center; position: absolute; left: 0; top: -2px; font-size: 14px; font-weight: bold; }}
        .action-btns {{ display: flex; gap: 20px; flex-wrap: wrap; }}
        .action-btns .btn {{ padding: 16px 40px; font-size: 18px; text-align: center; flex: 1; min-width: 200px; }}
        .btn-outline {{ background: transparent; border: 2px solid var(--gray-border); color: var(--ink); }}
        .btn-outline:hover {{ border-color: var(--charcoal); }}
        .specs-section {{ padding: 100px 5%; background: var(--gray-light); border-top: 1px solid var(--gray-border); }}
        .specs-container {{ max-width: 1000px; margin: 0 auto; background: var(--white); padding: 50px; border-radius: 16px; box-shadow: 0 10px 30px rgba(0,0,0,0.03); border: 1px solid var(--gray-border); }}
        .specs-title {{ font-size: 32px; color: var(--ink); margin-bottom: 40px; border-bottom: 3px solid var(--brand-yellow); padding-bottom: 15px; display: inline-block; }}
        .specs-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 20px; }}
        .spec-item {{ display: flex; justify-content: space-between; padding-bottom: 14px; border-bottom: 1px solid var(--gray-border); }}
        .spec-label {{ font-weight: 600; color: var(--charcoal); }}
        .spec-value {{ color: #777; font-weight: 500; text-align: right; }}
        footer {{ background-color: var(--ink); color: var(--white); padding: 80px 5% 40px; border-top: 5px solid var(--brand-yellow); }}
        .footer-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 40px; margin-bottom: 60px; }}
        .footer-col h4 {{ color: var(--brand-yellow); margin-bottom: 24px; font-size: 20px; text-transform: uppercase; }}
        .about-text {{ color: #A0AEC0; font-size: 15px; line-height: 1.8; margin-bottom: 20px; }}
        .footer-col ul {{ list-style: none; }}
        .footer-col ul li {{ margin-bottom: 12px; }}
        .footer-col ul a, .footer-col p {{ color: #A0AEC0; text-decoration: none; font-size: 14px; transition: color 0.2s; display: inline-block; line-height: 1.6; }}
        .footer-col ul a:hover {{ color: var(--white); }}
        .footer-bottom {{ text-align: center; color: #555; padding-top: 30px; border-top: 1px solid #333; font-size: 14px; }}
        @media (max-width: 900px) {{ .nav-links {{ display: none; }} .menu-toggle {{ display: block; }} .product-layout {{ padding: 40px 5%; gap: 40px; }} .product-title {{ font-size: 32px; }} }}
        @media (max-width: 600px) {{ .action-btns {{ flex-direction: column; }} .action-btns .btn {{ width: 100%; }} .specs-container {{ padding: 30px 20px; }} }}
</style>
</head>
<body>
{NAV_HTML}
<section class="product-layout">
<div class="product-gallery">
<div class="main-image"><img alt="{title}" src="{image}"/></div>
</div>
<div class="product-details">
<div class="product-category">{category}</div>
<h1 class="product-title">{title}</h1>
<div class="product-price">{price_display}</div>
<ul class="feature-list">
{features_html}
</ul>
<div class="action-btns">
<a class="btn" href="contact_v5_premium.html">Get Quote</a>
<a class="btn btn-outline" href="spare_parts_v1.html">Spare Parts</a>
</div>
</div>
</section>
<section class="specs-section">
<div class="specs-container">
<h2 class="specs-title">Technical Specifications</h2>
<div class="specs-grid">
{specs_html}
</div>
</div>
</section>
<footer>
<div class="footer-grid">
<div class="footer-col" style="grid-column: span 2;">
<h4>About Us</h4>
<p class="about-text">We are a manufacturer of Agriculture Equipment. We design, develop and MAKE IN INDIA.</p>
</div>
<div class="footer-col">
<h4>Quick Links</h4>
<ul>
<li><a href="product_catalog_v1.html">Products</a></li>
<li><a href="spare_parts_v1.html">Spare Parts</a></li>
<li><a href="about_v1.html">About</a></li>
<li><a href="contact_v5_premium.html">Contact</a></li>
</ul>
</div>
<div class="footer-col">
<h4>GET IN TOUCH</h4>
<p style="margin:0;">Email: info@krishitek.com</p>
<p style="margin:0;">Phone: +91 91570 62093</p>
<p style="margin:0;">Santej, Ahmedabad, Gujarat</p>
</div>
</div>
<div class="footer-bottom">&copy; 2026 KrishiTek. All rights reserved.</div>
</footer>
<script src="../assets/chatbot_widget.js"></script>
</body>
</html>'''


# ===== BUILD ALL PRODUCT PAGES =====
print("\n=== BUILDING PRODUCT PAGES ===")
for product_name, pmap in PRODUCT_MAP.items():
    if product_name in products:
        pdata = products[product_name]
        html = build_product_page(product_name, pdata, pmap)
        fpath = os.path.join(VARIANTS, pmap['file'])
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"  BUILT: {pmap['file']} ({len(html)} bytes) - {product_name}")
    else:
        print(f"  SKIP: {product_name} - no specs found in Excel")

# ===== APPLY NAV TO NON-PRODUCT PAGES =====
SKIP = {
    'homepage_v4_stitch.html',
    'dealer_login.html', 'dealer_dashboard.html', 'master_dashboard.html',
    'md_dashboard.html', 'employee_portal.html', 'login.html', 'ai_chatbot_dashboard.html',
    'homepage_v5_special.html', 'productpage_v1.html',
    'about_v2_uiux_promax.html', 'about_v4_practical.html',
    'contact_v3_uiux_promax.html', 'contact_v4_practical.html',
}
# Add all product pages we just built
for pmap in PRODUCT_MAP.values():
    SKIP.add(pmap['file'])
# Add spare pages too
for f in os.listdir(VARIANTS):
    if f.startswith('spare_') and f.endswith('_product_page.html'):
        SKIP.add(f)

print("\n=== APPLYING NAV TO REMAINING PAGES ===")
nav_count = 0
for fname in os.listdir(VARIANTS):
    if not fname.endswith('.html') or fname in SKIP:
        continue
    fpath = os.path.join(VARIANTS, fname)
    with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
        html = f.read()
    if len(html) < 100:
        continue
    original = html
    
    # Inject/Replace mega CSS
    css_pattern = re.compile(r'\s*/\* Mega Menu \*/.*?@media \(max-width: 900px\) { \.mega-menu { display: none !important; } }', re.DOTALL)
    if css_pattern.search(html):
        html = css_pattern.sub('\n' + MEGA_CSS.strip(), html)
    elif '</style>' in html:
        html = html.replace('</style>', MEGA_CSS + '\n    </style>', 1)
    
    # Add position relative to nav
    if 'position: relative' not in html.split('nav {')[0] if 'nav {' in html else html:
        html = html.replace('nav {', 'nav { position: relative;', 1)
    
    # Replace nav block
    nav_pattern = re.compile(r'<div class="utility-bar">.*?</nav>', re.DOTALL)
    if nav_pattern.search(html):
        html = nav_pattern.sub(NAV_HTML, html, count=1)
    
    # Fix links
    html = html.replace('href="contact_v2_stitch.html"', 'href="contact_v5_premium.html"')
    html = html.replace('href="blog_v2_stitch.html"', 'href="blog_v1.html"')
    html = html.replace('href="product_catalog_v2_stitch.html"', 'href="product_catalog_v1.html"')
    html = html.replace('href="spare_parts_v2_stitch.html"', 'href="spare_parts_v1.html"')
    
    if html != original:
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(html)
        nav_count += 1
        print(f"  NAV UPDATED: {fname}")

# Fix index.html
index_path = os.path.join(PROJ, 'index.html')
with open(index_path, 'w', encoding='utf-8') as f:
    f.write('<!DOCTYPE html>\n<html>\n<head><meta http-equiv="refresh" content="0; url=variants/homepage_v4_stitch.html"><title>KrishiTek</title></head>\n<body><p>Redirecting...</p></body>\n</html>\n')
print(f"\n  INDEX.HTML: Redirects to homepage_v4_stitch.html")

print(f"\n=== DONE: {len(PRODUCT_MAP)} product pages rebuilt, {nav_count} pages nav-updated ===")
