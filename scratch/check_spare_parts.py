import openpyxl

wb = openpyxl.load_workbook(r'KIPL_Product Specification_FM-DSN-05 (1)krishitek.xlsx', data_only=True)
with open('scratch/spare_parts_output.txt', 'w', encoding='utf-8') as f:
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        f.write(f"\n--- {sheetname} ---\n")
        for row in ws.iter_rows(values_only=True):
            row_str = ' | '.join(str(v) if v is not None else '' for v in row)
            if 'spare' in row_str.lower() or 'blade' in row_str.lower() or 'clutch' in row_str.lower():
                f.write(row_str + '\n')
